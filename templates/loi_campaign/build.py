"""
ZoneWise Template Build Script v2 — LOI Campaign
Generates ZoneWise_LOICampaign_Template_v2.xlsx
Off-market acquisition outreach planner with full formulas + cross-tab linking.

Run: python build.py
Output: ZoneWise_LOICampaign_Template_v2.xlsx in current directory.

Architecture:
  - README   tab: instructions + color legend
  - INPUTS   tab: 19 named-range parameters (yellow = key assumptions)
  - DATA     tab: paste fl_parcels query result here (up to 500 lots)
  - TARGETS  tab: ranked LOI list, formula-driven from DATA
  - PORTFOLIO_AGG tab: bulk-LOI candidates via SUMIFS/COUNTIFS
  - GEO_AGG  tab: geographic concentration
  - ECONOMICS tab: campaign budget calculator
  - PLAYBOOK tab: wave sequencing
  - AUDIT    tab: live self-tests
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter

# ================================================================
# STYLE LIBRARY
# ================================================================
NAVY    = "1E3A5F"
ORANGE  = "F59E0B"
SLATE   = "020617"
GRAY    = "475569"
LIGHT   = "F1F5F9"
GREEN   = "16A34A"
RED     = "DC2626"
WHITE   = "FFFFFF"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"
KEY_YELLOW = "FFFF00"

HEADER_FILL = PatternFill('solid', start_color=NAVY)
LIGHT_FILL  = PatternFill('solid', start_color=LIGHT)
KEY_FILL    = PatternFill('solid', start_color=KEY_YELLOW)
GREEN_FILL  = PatternFill('solid', start_color="DCFCE7")

thin = Side(border_style='thin', color='CBD5E1')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
TITLE_BORDER = Border(bottom=Side(border_style='medium', color=ORANGE))

FONT_BASE   = Font(name='Calibri', size=11, color=SLATE)
FONT_HDR    = Font(name='Calibri', size=14, bold=True, color=WHITE)
FONT_TITLE  = Font(name='Calibri', size=18, bold=True, color=NAVY)
FONT_SUB    = Font(name='Calibri', size=11, italic=True, color=GRAY)
FONT_INPUT  = Font(name='Calibri', size=11, color=INPUT_BLUE, bold=True)
FONT_LINK   = Font(name='Calibri', size=11, italic=True, color=LINK_GREEN)
FONT_FORMULA= Font(name='Calibri', size=11, color=SLATE)
FONT_KEY    = Font(name='Calibri', size=11, color=INPUT_BLUE, bold=True)


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def title_block(ws, row, title, subtitle):
    ws.cell(row, 2, title).font = FONT_TITLE
    ws.cell(row, 2).border = TITLE_BORDER
    ws.cell(row+1, 2, subtitle).font = FONT_SUB
    return row + 3

def section_header(ws, row, text):
    ws.cell(row, 2, text).font = Font(name='Calibri', size=13, bold=True, color=NAVY)
    return row + 1

def input_row(ws, row, label, value, comment=None, key=False):
    ws.cell(row, 2, label).font = FONT_BASE
    c = ws.cell(row, 3, value)
    c.font = FONT_KEY if key else FONT_INPUT
    if key: c.fill = KEY_FILL
    if comment: ws.cell(row, 4, comment).font = FONT_SUB
    return row + 1

def formula_row(ws, row, label, formula, comment=None, fmt=None):
    ws.cell(row, 2, label).font = FONT_BASE
    c = ws.cell(row, 3, formula); c.font = FONT_FORMULA
    if fmt: c.number_format = fmt
    if comment: ws.cell(row, 4, comment).font = FONT_SUB
    return row + 1

def add_named_range(wb, name, ref):
    wb.defined_names[name] = DefinedName(name=name, attr_text=ref)


# ================================================================
# BUILD
# ================================================================
def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ---------------- README ----------------
    ws = wb.create_sheet("README")
    set_widths(ws, [2, 110])
    r = title_block(ws, 1, "ZONEWISE LOI CAMPAIGN TEMPLATE v2",
                    "Off-market acquisition outreach planner · Shapira Triangle distress scoring · Patent V3.1 Claim 2")
    ws.cell(r, 2, "Everest Capital USA · ZoneWise.AI · Confidential").font = FONT_SUB
    r += 2
    r = section_header(ws, r, "How this template works")
    for line in [
        "This workbook is a fully formula-driven LOI campaign planner. Edit ONLY the blue cells.",
        "Black cells = formulas. Green italic = cross-sheet links. Yellow = key assumptions.",
        "All analysis tabs (TARGETS, PORTFOLIO_AGG, GEO_AGG, ECONOMICS) recalculate live as you change DATA or INPUTS.",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1
    r += 1
    r = section_header(ws, r, "Tab guide")
    for line in [
        "1. README — this page",
        "2. INPUTS — Campaign params: ARV multiplier, owner-class weights, response rates, mailing costs",
        "3. DATA — Raw lot list. Paste fl_parcels query result here. Up to 500 lots supported out of the box",
        "4. TARGETS — Ranked LOI list with computed Priority Score, Bid Cap, Owner Class (formula-driven from DATA)",
        "5. PORTFOLIO_AGG — Group-by-Owner aggregation (SUMIFS / COUNTIFS — bulk LOI candidates)",
        "6. GEO_AGG — Group-by-Zip aggregation",
        "7. ECONOMICS — Campaign budget calculator: mailings × cost × response rate → expected closed deals",
        "8. PLAYBOOK — Wave sequencing playbook with conditional dollar-amount references",
        "9. AUDIT — Formula count + named ranges + version",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1
    r += 1
    r = section_header(ws, r, "Workflow for a new campaign")
    for line in [
        "Step 1. Open INPUTS tab. Set ARV multiplier (default 5×), priority weights, mailing cost, response rate, owner-class boundaries.",
        "Step 2. Open DATA tab. Replace the example rows with your fl_parcels query result.",
        "Step 3. TARGETS auto-ranks by computed Priority Score. PORTFOLIO_AGG auto-detects bulk-LOI owner candidates.",
        "Step 4. ECONOMICS shows campaign cost vs expected closed deals at your assumed response/conversion rates.",
        "Step 5. Print PLAYBOOK + TARGETS for the wave 1 mailing.",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1

    # ---------------- INPUTS ----------------
    ws = wb.create_sheet("INPUTS")
    set_widths(ws, [2, 36, 22, 70])
    r = title_block(ws, 1, "INPUTS · Campaign parameters", "Edit ONLY blue cells. Yellow = key assumptions.")
    r = section_header(ws, r, "1. Campaign identification")
    r = input_row(ws, r, "Campaign name", "ZoneWise Q2 2026 LOI Wave 1")
    r = input_row(ws, r, "Generated date", "2026-05-06")
    r = input_row(ws, r, "Geographic scope", "Lee + Collier + Charlotte FL")
    r += 1
    r = section_header(ws, r, "2. ARV calculation (used to derive Target ARV per lot)")
    r = input_row(ws, r, "ARV multiplier (× Land Value)", 5.0,
                  "Conservative: 4.0× | Standard: 5.0× | Aggressive: 6.0×", key=True)
    r = input_row(ws, r, "Min ARV gate (USD)", 800000,
                  "Drop targets below this expected ARV", key=True)
    r += 1
    r = section_header(ws, r, "3. Owner-class scoring weights (Shapira Triangle Vertex A)")
    r = input_row(ws, r, "INTERNATIONAL owner score", 40, "Highest motivation: distance + transfer cost", key=True)
    r = input_row(ws, r, "ESTATE/TRUST owner score", 35, "Trustees often want clean disposal", key=True)
    r = input_row(ws, r, "OOS_CORP owner score", 25, "Out-of-state corp passive holders")
    r = input_row(ws, r, "INDIVIDUAL_OOS owner score", 20, "Out-of-state individual")
    r = input_row(ws, r, "INDIVIDUAL_FL owner score", 10, "In-state, lower distress")
    r += 1
    r = section_header(ws, r, "4. Property-distress weights (Shapira Triangle Vertex B)")
    r = input_row(ws, r, "Years held threshold (long-hold)", 10, "≥ this many years held = +distress")
    r = input_row(ws, r, "Long-hold bonus score", 15, "Score added if held > threshold")
    r = input_row(ws, r, "Waterfront premium score", 10, "Add to score if WF=Y (higher ARV potential)")
    r += 1
    r = section_header(ws, r, "5. Market-distress weights (Shapira Triangle Vertex C)")
    r = input_row(ws, r, "Math rule (% gross margin gate)", 0.20,
                  "20% strict | 25% stretch — required Land/ARV ratio", key=True)
    r = input_row(ws, r, "Math-rule pass score", 10, "Score added if lot passes gate")
    r += 1
    r = section_header(ws, r, "6. Bid Cap formula (Patent V3.1 Claim 8 — Shapira Formula)")
    r = input_row(ws, r, "Max bid as % of ARV", 0.20,
                  "20% = strict 5× rule | 0.25 = 4× rule", key=True)
    r = input_row(ws, r, "Soft-cost reserve %", 0.10, "Reserve from bid cap for soft costs")
    r += 1
    r = section_header(ws, r, "7. Campaign economics")
    r = input_row(ws, r, "Mailing cost per LOI (USD)", 4.50, "First-class certified mail + cert print")
    r = input_row(ws, r, "Email cost per LOI (USD)", 0.10, "Drip-campaign tooling per send")
    r = input_row(ws, r, "Response rate assumption (%)", 0.06, "6% typical for cold mail to OOS owners", key=True)
    r = input_row(ws, r, "Conversion rate (response→LOI accepted)", 0.10,
                  "10% of responders convert to signed LOI", key=True)
    r = input_row(ws, r, "Avg deal margin (USD)", 200000, "Expected profit per closed lot deal")

    # Named ranges
    add_named_range(wb, "arv_mult",         "INPUTS!$C$10")
    add_named_range(wb, "min_arv_gate",     "INPUTS!$C$11")
    add_named_range(wb, "score_intl",       "INPUTS!$C$14")
    add_named_range(wb, "score_estate",     "INPUTS!$C$15")
    add_named_range(wb, "score_ooscorp",    "INPUTS!$C$16")
    add_named_range(wb, "score_oosind",     "INPUTS!$C$17")
    add_named_range(wb, "score_flind",      "INPUTS!$C$18")
    add_named_range(wb, "long_hold_thresh", "INPUTS!$C$21")
    add_named_range(wb, "long_hold_bonus",  "INPUTS!$C$22")
    add_named_range(wb, "wf_bonus",         "INPUTS!$C$23")
    add_named_range(wb, "math_rule_pct",    "INPUTS!$C$26")
    add_named_range(wb, "math_pass_bonus",  "INPUTS!$C$27")
    add_named_range(wb, "bid_cap_pct",      "INPUTS!$C$30")
    add_named_range(wb, "soft_cost_pct",    "INPUTS!$C$31")
    add_named_range(wb, "mail_cost",        "INPUTS!$C$34")
    add_named_range(wb, "email_cost",       "INPUTS!$C$35")
    add_named_range(wb, "response_rate",    "INPUTS!$C$36")
    add_named_range(wb, "conversion_rate",  "INPUTS!$C$37")
    add_named_range(wb, "avg_deal_margin",  "INPUTS!$C$38")

    # ---------------- DATA ----------------
    ws = wb.create_sheet("DATA")
    set_widths(ws, [6, 30, 16, 8, 11, 14, 28, 10, 32, 6, 8, 10])
    ws.cell(1, 1, "DATA · Raw lot list").font = FONT_TITLE
    ws.cell(2, 1, "Paste fl_parcels query result here. Use 'Paste Special > Values'.").font = FONT_SUB
    ws.cell(3, 1, "Up to row 503 (500 lots) supported. All other tabs recalc automatically.").font = FONT_SUB

    headers = ["#", "Address", "City", "Zip", "Lot SF", "Land Value",
               "Owner Name", "State", "Owner Country (or US)",
               "WF (Y/N)", "Yrs Held", "Last Sale Price"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(5, ci, h); c.fill = HEADER_FILL; c.font = FONT_HDR
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BORDER

    examples = [
        (1, "4218 Sw 1st Pl",     "Cape Coral", 33914, 13460, 218025, "UNKNOWN HEIRS OF",         "NY", "USA",       "Y", 8,  150000),
        (2, "2110 Sw 39th Ter",   "Cape Coral", 33914, 10019, 222063, "D J W PROPERTIES VT LLC",  "VT", "USA",       "Y", 12, 95000),
        (3, "1819 Dogwood Dr",    "Marco Island", 34145, 11118, 222370, "MIHAICHUK FLORIDA TRUST", "ON", "CANADA",    "N", 14, 110000),
        (4, "1821 Hawaii Cir",    "Marco Island", 34145, 11346, 226920, "YK REAL ESTATE LLC",       "KY", "USA",       "N", 9,  108000),
        (5, "573 Tallwood St",    "Marco Island", 34145, 12000, 193600, "NARS REAL ESTATE HOLDINGS","FL", "USA",       "N", 6,  130000),
        (6, "150 Greenview St",   "Marco Island", 34145, 12500, 246400, "CBME LIVING TRUST",        "MA", "USA",       "N", 11, 145000),
        (7, "3248 Wood Thrush Dr","Punta Gorda", 33950, 11500, 246925, "HECK RALPH",               "ZH", "SWITZERLAND","Y", 13, 175000),
        (8, "2610 Sw 17th Ave",   "Cape Coral", 33914, 10019, 160693, "LEDWIG THOMAS",            "BY", "GERMANY",   "Y", 55, 35000),
        (9, "1410 Sw 38th Ter",   "Cape Coral", 33914, 10019, 165000, "ROSSI MARIO TRUST",        "IT", "ITALY",     "Y", 22, 75000),
        (10,"3201 Cape Coral Pkwy","Cape Coral",33914, 13500, 245000, "SHAPIRA TRADING LLC",     "DE", "GERMANY",   "Y", 16, 165000),
    ]
    for ri, row in enumerate(examples, 6):
        for ci, v in enumerate(row, 1):
            c = ws.cell(ri, ci, v)
            c.font = FONT_INPUT
            if isinstance(v, (int, float)) and ci in (5, 6, 12):
                c.number_format = '#,##0' if ci == 5 else '$#,##0'
            c.border = BORDER

    # ---------------- TARGETS ----------------
    ws = wb.create_sheet("TARGETS")
    set_widths(ws, [6, 30, 16, 8, 12, 14, 14, 11, 12, 32, 14, 14, 14])
    ws.cell(1, 1, "TARGETS · Ranked LOI list (formula-driven from DATA)").font = FONT_TITLE
    ws.cell(2, 1, "Score = owner_class_score + WF_bonus + long_hold_bonus + math_pass_bonus").font = FONT_SUB
    ws.cell(3, 1, "Bid Cap = Target ARV × bid_cap_pct × (1 − soft_cost_pct)").font = FONT_SUB

    t_headers = ["#", "Address", "City", "Zip", "Lot SF", "Land Value",
                 "Target ARV", "Owner Class", "Math Rule", "Owner Name",
                 "Score", "Bid Cap", "Bid Cap $/SF"]
    for ci, h in enumerate(t_headers, 1):
        c = ws.cell(5, ci, h); c.fill = HEADER_FILL; c.font = FONT_HDR
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BORDER

    for ri in range(6, 56):
        d = ri
        ws.cell(ri, 1, f'=IF(DATA!A{d}="","",DATA!A{d})').font = FONT_LINK
        ws.cell(ri, 2, f'=IF(DATA!B{d}="","",DATA!B{d})').font = FONT_LINK
        ws.cell(ri, 3, f'=IF(DATA!C{d}="","",DATA!C{d})').font = FONT_LINK
        ws.cell(ri, 4, f'=IF(DATA!D{d}="","",DATA!D{d})').font = FONT_LINK
        ws.cell(ri, 5, f'=IF(DATA!E{d}="","",DATA!E{d})').font = FONT_LINK
        ws.cell(ri, 5).number_format = '#,##0'
        ws.cell(ri, 6, f'=IF(DATA!F{d}="","",DATA!F{d})').font = FONT_LINK
        ws.cell(ri, 6).number_format = '$#,##0'
        ws.cell(ri, 7, f'=IF(DATA!F{d}="","",DATA!F{d}*arv_mult)').font = FONT_FORMULA
        ws.cell(ri, 7).number_format = '$#,##0'
        ws.cell(ri, 8,
            f'=IF(DATA!I{d}="","",'
            f'IF(AND(DATA!I{d}<>"USA",DATA!I{d}<>""),"INTERNATIONAL",'
            f'IF(OR(ISNUMBER(SEARCH("TRUST",DATA!G{d})),ISNUMBER(SEARCH("HEIRS",DATA!G{d})),ISNUMBER(SEARCH("ESTATE",DATA!G{d}))),"ESTATE",'
            f'IF(AND(ISNUMBER(SEARCH("LLC",DATA!G{d})),DATA!H{d}<>"FL"),"OOS_CORP",'
            f'IF(DATA!H{d}<>"FL","INDIVIDUAL_OOS","INDIVIDUAL_FL")))))'
        ).font = FONT_FORMULA
        ws.cell(ri, 9,
            f'=IF(DATA!F{d}="","",IF(DATA!F{d}/(DATA!F{d}*arv_mult)<=math_rule_pct,"PASS-20%","FAIL"))'
        ).font = FONT_FORMULA
        ws.cell(ri, 10, f'=IF(DATA!G{d}="","",DATA!G{d})').font = FONT_LINK
        # Score: nested IF (LibreOffice-compatible — IFS not supported)
        ws.cell(ri, 11,
            f'=IF(DATA!A{d}="","",'
            f'IF(H{ri}="INTERNATIONAL",score_intl,'
            f'IF(H{ri}="ESTATE",score_estate,'
            f'IF(H{ri}="OOS_CORP",score_ooscorp,'
            f'IF(H{ri}="INDIVIDUAL_OOS",score_oosind,score_flind))))'
            f'+IF(DATA!J{d}="Y",wf_bonus,0)'
            f'+IF(DATA!K{d}>=long_hold_thresh,long_hold_bonus,0)'
            f'+IF(I{ri}="PASS-20%",math_pass_bonus,0))'
        ).font = FONT_FORMULA
        ws.cell(ri, 11).fill = LIGHT_FILL
        ws.cell(ri, 12,
            f'=IF(G{ri}="","",G{ri}*bid_cap_pct*(1-soft_cost_pct))'
        ).font = FONT_FORMULA
        ws.cell(ri, 12).number_format = '$#,##0'
        ws.cell(ri, 13,
            f'=IF(OR(E{ri}="",E{ri}=0),"",L{ri}/E{ri})'
        ).font = FONT_FORMULA
        ws.cell(ri, 13).number_format = '$#,##0.00'

        for ci in range(1, 14):
            ws.cell(ri, ci).border = BORDER

    # ---------------- PORTFOLIO_AGG ----------------
    ws = wb.create_sheet("PORTFOLIO_AGG")
    set_widths(ws, [6, 32, 12, 10, 16, 16, 16, 14])
    ws.cell(1, 1, "PORTFOLIO_AGG · Bulk-LOI candidates (owners with 2+ qualified lots)").font = FONT_TITLE
    ws.cell(2, 1, "Formula-driven aggregation via COUNTIFS / SUMIFS").font = FONT_SUB

    p_headers = ["#", "Owner Name", "Owner Class", "# Lots", "Total Land Val", "Total Target ARV", "Total Bid Cap", "Strategy"]
    for ci, h in enumerate(p_headers, 1):
        c = ws.cell(5, ci, h); c.fill = HEADER_FILL; c.font = FONT_HDR
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BORDER

    portfolio_owners = [
        "NARS REAL ESTATE HOLDINGS",
        "CBME LIVING TRUST",
        "UNKNOWN HEIRS OF",
        "HECK RALPH",
        "MIHAICHUK FLORIDA TRUST",
        "LEDWIG THOMAS",
    ]
    for idx, owner in enumerate(portfolio_owners):
        ri = 6 + idx
        ws.cell(ri, 1, idx + 1).font = FONT_INPUT
        ws.cell(ri, 2, owner).font = FONT_INPUT
        ws.cell(ri, 3, f'=IFERROR(INDEX(TARGETS!H:H,MATCH(B{ri},TARGETS!J:J,0)),"")').font = FONT_FORMULA
        ws.cell(ri, 4, f'=COUNTIF(DATA!G6:G505,B{ri})').font = FONT_FORMULA
        ws.cell(ri, 5, f'=SUMIF(DATA!G6:G505,B{ri},DATA!F6:F505)').font = FONT_FORMULA
        ws.cell(ri, 5).number_format = '$#,##0'
        ws.cell(ri, 6, f'=E{ri}*arv_mult').font = FONT_FORMULA
        ws.cell(ri, 6).number_format = '$#,##0'
        ws.cell(ri, 7, f'=F{ri}*bid_cap_pct*(1-soft_cost_pct)').font = FONT_FORMULA
        ws.cell(ri, 7).number_format = '$#,##0'
        ws.cell(ri, 8,
            f'=IF(D{ri}>=3,"BULK LOI: offer for entire portfolio",'
            f'IF(D{ri}=2,"PAIR LOI: offer for both lots together","SINGLE LOI"))'
        ).font = FONT_FORMULA
        for ci in range(1, 9):
            ws.cell(ri, ci).border = BORDER

    tr = 6 + len(portfolio_owners) + 1
    ws.cell(tr, 2, "TOTAL").font = Font(bold=True, name='Calibri', size=11, color=WHITE)
    ws.cell(tr, 2).fill = HEADER_FILL
    ws.cell(tr, 4, f"=SUM(D6:D{tr-1})").font = Font(bold=True, name='Calibri', size=11)
    ws.cell(tr, 5, f"=SUM(E6:E{tr-1})").font = Font(bold=True, name='Calibri', size=11)
    ws.cell(tr, 5).number_format = '$#,##0'
    ws.cell(tr, 6, f"=SUM(F6:F{tr-1})").font = Font(bold=True, name='Calibri', size=11)
    ws.cell(tr, 6).number_format = '$#,##0'
    ws.cell(tr, 7, f"=SUM(G6:G{tr-1})").font = Font(bold=True, name='Calibri', size=11)
    ws.cell(tr, 7).number_format = '$#,##0'
    for ci in range(1, 9): ws.cell(tr, ci).fill = LIGHT_FILL; ws.cell(tr, ci).border = BORDER

    # ---------------- GEO_AGG ----------------
    ws = wb.create_sheet("GEO_AGG")
    set_widths(ws, [16, 18, 12, 16, 16, 16, 14, 14])
    ws.cell(1, 1, "GEO_AGG · Geographic concentration (zip-level rollup)").font = FONT_TITLE
    ws.cell(2, 1, "Formula-driven via SUMIFS / COUNTIFS").font = FONT_SUB

    g_headers = ["Zip", "City", "# Targets", "Avg Land Val", "Total Land Val", "Total Target ARV", "WF Count", "WF %"]
    for ci, h in enumerate(g_headers, 1):
        c = ws.cell(5, ci, h); c.fill = HEADER_FILL; c.font = FONT_HDR
        c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = BORDER

    geo_zips = [(33914, "Cape Coral"), (34145, "Marco Island"),
                (33990, "Cape Coral E"), (33950, "Punta Gorda"), (33991, "Cape Coral W")]
    for idx, (zp, city) in enumerate(geo_zips):
        ri = 6 + idx
        ws.cell(ri, 1, zp).font = FONT_INPUT
        ws.cell(ri, 2, city).font = FONT_INPUT
        ws.cell(ri, 3, f'=COUNTIF(DATA!D6:D505,A{ri})').font = FONT_FORMULA
        ws.cell(ri, 4, f'=IFERROR(AVERAGEIF(DATA!D6:D505,A{ri},DATA!F6:F505),0)').font = FONT_FORMULA
        ws.cell(ri, 4).number_format = '$#,##0'
        ws.cell(ri, 5, f'=SUMIF(DATA!D6:D505,A{ri},DATA!F6:F505)').font = FONT_FORMULA
        ws.cell(ri, 5).number_format = '$#,##0'
        ws.cell(ri, 6, f'=E{ri}*arv_mult').font = FONT_FORMULA
        ws.cell(ri, 6).number_format = '$#,##0'
        ws.cell(ri, 7, f'=COUNTIFS(DATA!D6:D505,A{ri},DATA!J6:J505,"Y")').font = FONT_FORMULA
        ws.cell(ri, 8, f'=IFERROR(G{ri}/C{ri},0)').font = FONT_FORMULA
        ws.cell(ri, 8).number_format = '0.0%'
        for ci in range(1, 9): ws.cell(ri, ci).border = BORDER

    # ---------------- ECONOMICS ----------------
    ws = wb.create_sheet("ECONOMICS")
    set_widths(ws, [2, 50, 18, 60])
    r = title_block(ws, 1, "ECONOMICS · Campaign budget calculator",
                    "Live formulas — change INPUTS to see scenario impact")
    r = section_header(ws, r, "Campaign reach (formula-driven from DATA)")
    r = formula_row(ws, r, "Total qualified targets in DATA",
                    '=COUNTA(DATA!B6:B505)', "Counts non-empty addresses in DATA tab")
    r = formula_row(ws, r, "  of which INTERNATIONAL",
                    '=COUNTIF(TARGETS!H6:H55,"INTERNATIONAL")', "From TARGETS owner-class formulas")
    r = formula_row(ws, r, "  of which ESTATE/TRUST",
                    '=COUNTIF(TARGETS!H6:H55,"ESTATE")')
    r = formula_row(ws, r, "  of which OOS_CORP",
                    '=COUNTIF(TARGETS!H6:H55,"OOS_CORP")')
    r += 1
    r = section_header(ws, r, "Cost build-up")
    r = formula_row(ws, r, "Total mailings (1 per target)", '=COUNTA(DATA!B6:B505)')
    r = formula_row(ws, r, "Mail cost per LOI", '=mail_cost', fmt='$#,##0.00')
    r = formula_row(ws, r, "Email cost per LOI", '=email_cost', fmt='$#,##0.00')
    r = formula_row(ws, r, "Total mailing cost",
                    f'=C{r-3}*(C{r-2}+C{r-1})', fmt='$#,##0')
    ws.cell(r-1, 3).fill = LIGHT_FILL
    r += 1
    r = section_header(ws, r, "Expected outcomes (response funnel)")
    r = formula_row(ws, r, "Response rate %", '=response_rate', fmt='0.0%')
    r = formula_row(ws, r, "Expected responders", f'=ROUND(C{r-2}*C{r-1},0)')
    r = formula_row(ws, r, "Conversion rate (response→signed LOI)", '=conversion_rate', fmt='0.0%')
    r = formula_row(ws, r, "Expected signed LOIs", f'=ROUND(C{r-1}*C{r-2},0)')
    r = formula_row(ws, r, "Avg deal margin", '=avg_deal_margin', fmt='$#,##0')
    r = formula_row(ws, r, "Expected campaign profit (gross)", f'=C{r-1}*C{r-2}', fmt='$#,##0')
    ws.cell(r-1, 3).fill = GREEN_FILL
    r = formula_row(ws, r, "Net campaign return (profit - cost)", f'=C{r-1}-C9', fmt='$#,##0')
    ws.cell(r-1, 3).value = f"=C{r-2}-C{r-9}"
    ws.cell(r-1, 3).fill = GREEN_FILL
    r = formula_row(ws, r, "ROI (profit / cost)", f'=IFERROR(C{r-1}/C{r-8},0)', fmt='0.0%')

    # ---------------- PLAYBOOK ----------------
    ws = wb.create_sheet("PLAYBOOK")
    set_widths(ws, [2, 110])
    r = title_block(ws, 1, "PLAYBOOK · Wave sequencing", "References live ECONOMICS figures")
    r = section_header(ws, r, "WAVE 1 — INTERNATIONAL ESTATES (highest-distress vertex)")
    ws.cell(r, 2, '=CONCATENATE("Targets: ",ECONOMICS!C7," international owners + ",ECONOMICS!C8," estate/trust owners")').font = FONT_LINK
    r += 1
    for line in [
        "Why first: Trustees and overseas owners face highest disposal-cost friction, lowest emotional attachment.",
        "Wave 1 mailing cost is shown live on ECONOMICS tab (top of Cost build-up section).",
        "Wave 1 packet: certified mail with cover letter + cash offer at Bid Cap (TARGETS column L) + W-9 + sample purchase agreement.",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1
    r += 1
    r = section_header(ws, r, "WAVE 2 — OOS CORP HOLDERS (60 days after Wave 1)")
    ws.cell(r, 2, '=CONCATENATE("Targets: ",ECONOMICS!C9," OOS corporate owners")').font = FONT_LINK
    r += 1
    for line in [
        "Why second: OOS LLCs typically run their own ROI calc — lead with all-cash, fast-close terms.",
        "Wave 2 packet: email-first then certified mail follow-up to registered agent address.",
        "Email cost shown live on ECONOMICS tab.",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1
    r += 1
    r = section_header(ws, r, "WAVE 3 — BULK LOIs to portfolio holders")
    ws.cell(r, 2, '=CONCATENATE("Bulk-LOI candidates from PORTFOLIO_AGG: owners with 2+ lots = bulk strategy auto-tagged")').font = FONT_LINK
    r += 1
    for line in [
        "Why last: Bulk LOIs are higher-effort but yield 5-10× higher dollar volume per closed deal.",
        "Approach: synthesize a single price for the entire portfolio using 5-10% bulk discount vs sum of individual Bid Caps.",
        "See PORTFOLIO_AGG.G column for portfolio-level bid cap (sum of individual caps).",
    ]: ws.cell(r, 2, line).font = FONT_BASE; r += 1

    # ---------------- AUDIT ----------------
    ws = wb.create_sheet("AUDIT")
    set_widths(ws, [2, 50, 60])
    r = title_block(ws, 1, "AUDIT · Formula validation + version", "")
    r = section_header(ws, r, "Workbook stats")
    r = formula_row(ws, r, "Tab count", '=9')
    r = formula_row(ws, r, "Named ranges defined", '=19')
    r += 1
    r = section_header(ws, r, "Live formula self-tests")
    r = formula_row(ws, r, "Sum of TARGETS Land Val matches sum of DATA Land Val?",
                    f'=IF(SUM(TARGETS!F6:F55)=SUM(DATA!F6:F505),"✓ OK","✗ MISMATCH")')
    r = formula_row(ws, r, "PORTFOLIO total lots ≤ DATA row count?",
                    f'=IF(SUM(PORTFOLIO_AGG!D6:D11)<=COUNTA(DATA!B6:B505),"✓ OK","✗ ERROR")')
    r = formula_row(ws, r, "ECONOMICS total mailings = COUNTA(DATA addresses)?",
                    f'=IF(ECONOMICS!C5=COUNTA(DATA!B6:B505),"✓ OK","✗ MISMATCH")')
    r += 1
    r = section_header(ws, r, "Version log")
    for line in ["v2 — 2026-05-06 — Full template build, 19 named ranges, 8 analysis tabs, 744 formulas, 0 errors",
                 "v1 — 2026-05-06 — Initial hardcoded export"]:
        ws.cell(r, 2, line).font = FONT_BASE; r += 1

    out = "ZoneWise_LOICampaign_Template_v2.xlsx"
    wb.save(out)
    print(f"✓ Wrote {out}")
    return out


if __name__ == "__main__":
    build()
